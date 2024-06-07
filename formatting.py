import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, Color, PatternFill

thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'))

bold_font = Font(bold=True)

gray_fill = PatternFill(start_color='D2D2D2', fill_type='solid')

orange_fill = PatternFill(start_color='FFC163', fill_type='solid')

large_font = Font(size=40, bold=True)
