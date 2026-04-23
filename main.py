import datetime
from openpyxl import Workbook
from methods import *
from customer import Customer
from meter import Meter
from formatting import *


def main():
    logging.info("Process Started")

    path = ("C:/Users/gduke/OneDrive - City of Georgetown/Documents - Team - Water Utility/Utility Support/Analytics/Products/Large Property Variance/")
    output_path = path + 'Results/'

    # Instantiate data df
    data = init_df(path + 'sample_book.xlsx')

    # Perform date calc, get calculated monday from previous week
    today = datetime.date.today()
    date = (pd.Timestamp(today) - pd.Timedelta(days=pd.Timestamp(today).dayofweek + 7)).date()

    # Instantiate dcp manually
    dcp = 1

    customers = []
    sources, vars = [], []
    for row in data.itertuples(index=True, name='Customer'):
        # Create customer obj and set object variables from data file
        customer = Customer(name=row.CustomerName, footage=row.IrrigatableArea, cust_type=row.Type)
        customer.set_allowance(allowance=calculate_budget(row.IrrigatableArea, dcp))
        for item in [x.strip() for x in str(int(row.SourceID)).split(",") if x.strip()]:
            sources.append(item)
        for item in [x.strip() for x in str(int(row.VariableID)).split(",") if x.strip()]:
            vars.append(item)
        for count, item in enumerate([x.strip() for x in str(row.IrrigationMeters).split(",") if x.strip()]):
            customer.add_meter(Meter(item, 'Irrigation', sources[count], vars[count]))

        for meter in customer.meters:
            if customer.type == 'MDM':
                meter_data = query_mdm_intervals(meter, date)
            else:
                meter_data = query_dh(meter.sourceid, date, meter.variableid)
                meter.add_data(meter_data)

            customer.add_usage(meter_data.ReadValue.sum())
            customer.mon_viol = monday_violations(meter_data)
            # Calculate midday or Monday violations given DCP
            if dcp < 3:
                customer.add_days(midday_violations(meter_data, dcp))
            elif dcp >= 3:
                customer.irrig_viol += irrigation_violations(meter_data)

        # If total usage from all meters exceeds customer budget, add violation
        if dcp < 3:
            if customer.usage > customer.allowance:
                customer.bug_viol = 1

            days = [day for sublist in customer.mid_days for day in sublist]
            print(days)
            customer.mid_viol = len(set(days))

        customers.append(customer)
        logging.info('Processed', customer.name, 'Allowance:', customer.allowance, 'Usage:', customer.usage)
        print(customer.name, customer.meters)

    # Create workbook
    wb = Workbook()
    ws = wb.active

    # Manually set column widths of output file
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['L'].width = 10

    # Manually set column names,
    ws.cell(row=1, column=1).value = 'Customer Name'
    ws.cell(row=1, column=1).border = thin_border
    ws.cell(row=1, column=1).font = bold_font
    ws.cell(row=1, column=2).value = 'Customer Number'
    ws.cell(row=1, column=2).border = thin_border
    ws.cell(row=1, column=2).font = bold_font
    ws.cell(row=1, column=3).value = 'Budget Violation'
    ws.cell(row=1, column=3).border = thin_border
    ws.cell(row=1, column=3).font = bold_font
    ws.cell(row=1, column=4).value = 'Irrigation Violations'
    ws.cell(row=1, column=4).border = thin_border
    ws.cell(row=1, column=4).font = bold_font
    ws.cell(row=1, column=5).value = 'Mid-day Violations'
    ws.cell(row=1, column=5).border = thin_border
    ws.cell(row=1, column=5).font = bold_font
    ws.cell(row=1, column=6).value = 'Monday Violations'
    ws.cell(row=1, column=6).border = thin_border
    ws.cell(row=1, column=6).font = bold_font
    ws.cell(row=1, column=7).value = 'Customer Budget'
    ws.cell(row=1, column=7).border = thin_border
    ws.cell(row=1, column=7).font = bold_font
    ws.cell(row=1, column=8).value = 'Customer Usage'
    ws.cell(row=1, column=8).border = thin_border
    ws.cell(row=1, column=8).font = bold_font
    ws['J1'] = 'Week Of'
    ws['J1'].border = thin_border
    ws['J1'].font = bold_font
    ws['J2'] = date
    ws['L1'] = 'DCP Stage'
    ws['L1'].border = thin_border
    ws['L1'].font = bold_font
    ws['L2'] = dcp

    for i, customer in enumerate(customers):
        ws.cell(row=i + 2, column=1).value = customer.name
        ws.cell(row=i + 2, column=2).value = customer.get_acc_party()
        ws.cell(row=i + 2, column=3).value = customer.bug_viol
        ws.cell(row=i + 2, column=4).value = customer.irrig_viol
        ws.cell(row=i + 2, column=5).value = customer.mid_viol
        ws.cell(row=i + 2, column=6).value = customer.mon_viol
        ws.cell(row=i + 2, column=7).value = customer.allowance
        ws.cell(row=i + 2, column=8).value = customer.usage

    wb.save(output_path + 'output.xlsx')

    logging.info("Process completed\n" + "=" * 50)


if __name__ == "__main__":
    main()
